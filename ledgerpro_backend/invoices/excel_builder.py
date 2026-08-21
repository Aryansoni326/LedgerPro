"""Shared Excel workbook builder for invoice exports."""
from __future__ import annotations

import io
from collections import defaultdict
from collections.abc import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Bill

COLUMNS = [
    'Date of Bill', 'Invoice Number', 'Seller (From)', 'Seller GSTIN',
    'Buyer (To)', 'Buyer GSTIN', 'Place of Supply', 'Taxable Amount',
    'CGST', 'SGST', 'IGST', 'Total Amount',
]


def _sheet_rows(bills_sub: Iterable[Bill]) -> list[dict]:
    rows = []
    for bill in bills_sub:
        raw = bill.raw_data or {}
        rows.append({
            'Date of Bill': raw.get('invoice_date', ''),
            'Invoice Number': raw.get('invoice_number', ''),
            'Seller (From)': raw.get('party_name_from', ''),
            'Seller GSTIN': raw.get('gstin_from', ''),
            'Buyer (To)': raw.get('party_name_to', ''),
            'Buyer GSTIN': raw.get('gstin_to', ''),
            'Place of Supply': raw.get('place_of_supply', ''),
            'Taxable Amount': float(raw.get('taxable_amount', 0.0) or 0.0),
            'CGST': float(raw.get('cgst', 0.0) or 0.0),
            'SGST': float(raw.get('sgst', 0.0) or 0.0),
            'IGST': float(raw.get('igst', 0.0) or 0.0),
            'Total Amount': float(raw.get('total_amount', 0.0) or 0.0),
        })
    return rows


def build_invoice_excel_bytes(bills_list: list[Bill]) -> bytes:
    """Build a styled multi-sheet workbook grouped by upload date."""
    bills_by_date = defaultdict(list)
    for bill in bills_list:
        bills_by_date[bill.uploaded_at.date()].append(bill)

    sorted_dates = sorted(bills_by_date.keys())
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for day in sorted_dates:
            sheet_name = day.strftime('%d-%m-%Y')
            df = pd.DataFrame(_sheet_rows(bills_by_date[day]), columns=COLUMNS)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for day in sorted_dates:
            sheet_name = day.strftime('%d-%m-%Y')
            worksheet = workbook[sheet_name]
            worksheet.freeze_panes = 'A2'

            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_idx in range(1, 13):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 13)

            currency_format = '"₹"#,##0.00'
            for row in range(2, worksheet.max_row + 1):
                for col_idx in [8, 9, 10, 11, 12]:
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = currency_format
                    cell.alignment = Alignment(horizontal='right')

            total_row_idx = worksheet.max_row + 1
            total_font = Font(name='Calibri', size=11, bold=True)
            worksheet.cell(row=total_row_idx, column=1, value='Total').font = total_font

            border_style = Border(
                top=Side(style='thin', color='A0A0A0'),
                bottom=Side(style='double', color='1F4E79'),
            )

            for col_idx in [8, 9, 10, 11, 12]:
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
                cell = worksheet.cell(row=total_row_idx, column=col_idx, value=formula)
                cell.font = total_font
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right')
                cell.border = border_style

    return output.getvalue()


def rebuild_excel_bytes_for_batch(batch) -> bytes:
    """
    Rebuild workbook bytes from a batch's stored bill_ids.
    Used when the original media file was lost (e.g. ephemeral disk).
    """
    bill_ids = batch.bill_ids or []
    bills_list = list(
        Bill.objects.filter(firm=batch.firm, id__in=bill_ids, is_deleted=False)
        .order_by('-uploaded_at')
    )
    if not bills_list:
        # Include soft-deleted source rows if needed so historical exports still download.
        bills_list = list(
            Bill.objects.filter(firm=batch.firm, id__in=bill_ids).order_by('-uploaded_at')
        )
    if not bills_list:
        raise FileNotFoundError('No invoice rows remain to rebuild this Excel export.')
    return build_invoice_excel_bytes(bills_list)
