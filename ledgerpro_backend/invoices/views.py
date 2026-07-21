import logging
import re
from io import BytesIO

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from audit.models import AuditLog
from audit.services import log_audit
from common.upload_validation import (
    DOCUMENT_EXTENSIONS,
    UploadValidationError,
    validate_upload,
)
from firms.access import get_firm_or_403
from firms.permissions import HasFirmAccess

from .models import Bill, ExcelExportBatch
from .services import InvoiceStorageService
from .tasks import extract_invoice_data

logger = logging.getLogger(__name__)

@api_view(['POST'])
@permission_classes([IsAuthenticated, HasFirmAccess])
def upload_invoices(request, firm_id):
    """
    POST: Upload multiple invoices/bills (PDF, JPG, PNG) for a specific firm.
    Validates file format and size (<10MB), creates database records with 'processing' status,
    and enqueues the Celery extraction task.
    """
    firm = get_firm_or_403(request, firm_id)
    if isinstance(firm, Response):
        return firm

    files = request.FILES.getlist('files') or request.FILES.getlist('file')
    if not files:
        return Response({'error': 'No files provided for upload.'}, status=status.HTTP_400_BAD_REQUEST)

    uploaded_bills = []
    errors = []

    for f in files:
        try:
            file_bytes, _ext = validate_upload(f, allowed_extensions=DOCUMENT_EXTENSIONS)
        except UploadValidationError as exc:
            errors.append(f"File '{f.name}': {exc}")
            continue

        try:
            buffer = BytesIO(file_bytes)
            buffer.name = f.name
            file_url = InvoiceStorageService.upload_invoice(buffer, firm.id)

            bill = Bill.objects.create(
                firm=firm,
                file_name=f.name,
                file_url=file_url,
                file_size=len(file_bytes),
                status='processing',
                uploaded_by=request.user
            )

            from vault.models import CloudVaultEntry
            CloudVaultEntry.objects.create(
                firm=firm,
                bill=bill,
                file_name=bill.file_name,
                file_url=bill.file_url,
                module='invoices',
                is_finalized=False
            )

            extract_invoice_data.delay(bill.id)

            log_audit(
                user=request.user,
                firm=firm,
                resource_type=AuditLog.RESOURCE_BILL,
                resource_id=bill.id,
                action=AuditLog.ACTION_UPLOAD,
                details={'file_name': bill.file_name, 'file_size': bill.file_size},
                request=request,
            )

            uploaded_bills.append({
                'id': bill.id,
                'file_name': bill.file_name,
                'file_url': bill.file_url,
                'status': bill.status,
                'file_size': bill.file_size
            })

        except Exception as e:
            logger.error("Failed to process file upload for %s: %s", f.name, e, exc_info=True)
            errors.append(f"Failed to process file upload for '{f.name}'.")

    if not uploaded_bills and errors:
        return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'uploaded': uploaded_bills,
        'errors': errors if errors else None
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasFirmAccess])
def list_invoices(request, firm_id):
    """
    GET: List all invoices/bills associated with a specific client firm.
    Optional query parameter ?status=processing|needs_review|approved
    """
    firm = get_firm_or_403(request, firm_id)
    if isinstance(firm, Response):
        return firm

    bills = Bill.objects.filter(firm=firm, is_deleted=False).order_by('-uploaded_at')

    # Apply filters
    status_filter = request.query_params.get('status')
    if status_filter:
        bills = bills.filter(status=status_filter)

    bill_type_filter = request.query_params.get('bill_type')
    if bill_type_filter:
        # Check raw_data JSON field bill_type
        bills = bills.filter(raw_data__bill_type=bill_type_filter)

    start_date = request.query_params.get('start_date')
    if start_date:
        bills = bills.filter(uploaded_at__date__gte=start_date)

    end_date = request.query_params.get('end_date')
    if end_date:
        bills = bills.filter(uploaded_at__date__lte=end_date)

    search = request.query_params.get('search')
    if search:
        # Search by party_name or invoice_number or file_name
        from django.db.models import Q
        bills = bills.filter(
            Q(file_name__icontains=search) |
            Q(raw_data__party_name__icontains=search) |
            Q(raw_data__invoice_number__icontains=search)
        )

    bills_data = [{
        'id': b.id,
        'file_name': b.file_name,
        'file_url': b.file_url,
        'status': b.status,
        'file_size': b.file_size,
        'uploaded_at': b.uploaded_at,
        'uploaded_by': b.uploaded_by.email,
        'raw_data': b.raw_data,
        'validation_warnings': b.validation_warnings,
        'extraction_failed': b.extraction_failed
    } for b in bills]

    return Response(bills_data, status=status.HTTP_200_OK)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def manage_invoice(request, pk):
    """
    PATCH: Update parsed invoice values inside raw_data and re-run validators.
    DELETE: Soft-delete the invoice (Bill) by setting is_deleted = True,
    and remove file references from any un-finalized CloudVaultEntries.
    """
    try:
        bill = Bill.objects.select_related('firm').get(pk=pk, is_deleted=False)
    except Bill.DoesNotExist:
        return Response({'error': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)

    permission = HasFirmAccess()
    if not permission.has_object_permission(request, None, bill.firm):
        return Response({'error': 'You do not have access to this invoice workspace.'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'DELETE':
        bill.is_deleted = True
        bill.save()

        from vault.models import CloudVaultEntry
        CloudVaultEntry.objects.filter(bill=bill, is_finalized=False).update(bill=None)

        log_audit(
            user=request.user,
            firm=bill.firm,
            resource_type=AuditLog.RESOURCE_BILL,
            resource_id=bill.id,
            action=AuditLog.ACTION_DELETE,
            details={'file_name': bill.file_name},
            request=request,
        )

        return Response({'message': 'Invoice soft-deleted successfully.'}, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        new_raw_data = request.data.get('raw_data')
        if not isinstance(new_raw_data, dict):
            return Response({'error': 'raw_data dictionary is required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Merge patch fields
        current_raw_data = bill.raw_data or {}
        current_raw_data.update(new_raw_data)
        bill.raw_data = current_raw_data

        # Re-run Server-Side Validation Rules Pass
        warnings = []

        taxable = float(current_raw_data.get('taxable_amount', 0.0) or 0.0)
        cgst = float(current_raw_data.get('cgst', 0.0) or 0.0)
        sgst = float(current_raw_data.get('sgst', 0.0) or 0.0)
        igst = float(current_raw_data.get('igst', 0.0) or 0.0)
        cess = float(current_raw_data.get('cess', 0.0) or 0.0)
        total = float(current_raw_data.get('total_amount', 0.0) or 0.0)

        # Arithmetic mismatch check
        computed_total = taxable + cgst + sgst + igst + cess
        if abs(computed_total - total) > 1.0:
            warnings.append(f"Total mismatch: taxable + taxes sum to ₹{computed_total:.2f}, but total_amount is ₹{total:.2f}.")

        # GSTIN check
        gstin_regex = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'
        gstin_from = current_raw_data.get('gstin_from')
        if gstin_from:
            if not re.match(gstin_regex, str(gstin_from).strip().upper()):
                warnings.append(f"Invalid supplier (From) GSTIN format: '{gstin_from}'.")
        gstin_to = current_raw_data.get('gstin_to')
        if gstin_to:
            if not re.match(gstin_regex, str(gstin_to).strip().upper()):
                warnings.append(f"Invalid buyer (To) GSTIN format: '{gstin_to}'.")

        # Intra vs Inter state checks
        has_intra = (cgst > 0 or sgst > 0)
        has_inter = (igst > 0)
        if (has_intra or has_inter):
            intra_valid = (cgst > 0 and sgst > 0 and igst == 0)
            inter_valid = (igst > 0 and cgst == 0 and sgst == 0)
            if not (intra_valid or inter_valid):
                warnings.append("Tax classification warning: conflicting GST structure (both CGST/SGST and IGST are populated, or CGST/SGST splits are uneven).")

        bill.validation_warnings = warnings
        bill.extraction_failed = False
        bill.save()

        log_audit(
            user=request.user,
            firm=bill.firm,
            resource_type=AuditLog.RESOURCE_BILL,
            resource_id=bill.id,
            action=AuditLog.ACTION_EDIT,
            details={'fields': list(new_raw_data.keys())},
            request=request,
        )

        return Response({
            'id': bill.id,
            'status': bill.status,
            'raw_data': bill.raw_data,
            'validation_warnings': bill.validation_warnings,
            'extraction_failed': bill.extraction_failed
        }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def verify_invoice(request, pk):
    """
    POST: Mark the invoice status as 'verified' after human confirmation.
    """
    try:
        bill = Bill.objects.select_related('firm').get(pk=pk, is_deleted=False)
    except Bill.DoesNotExist:
        return Response({'error': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)

    permission = HasFirmAccess()
    if not permission.has_object_permission(request, None, bill.firm):
        return Response({'error': 'You do not have access to this invoice workspace.'}, status=status.HTTP_403_FORBIDDEN)

    bill.status = 'verified'
    bill.save()

    log_audit(
        user=request.user,
        firm=bill.firm,
        resource_type=AuditLog.RESOURCE_BILL,
        resource_id=bill.id,
        action=AuditLog.ACTION_VERIFY,
        request=request,
    )

    return Response({
        'id': bill.id,
        'status': bill.status,
        'message': 'Invoice successfully verified.'
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def retry_extraction(request, pk):
    """
    POST: Reset status to 'processing' and retry Gemini parsing logic.
    """
    try:
        bill = Bill.objects.select_related('firm').get(pk=pk, is_deleted=False)
    except Bill.DoesNotExist:
        return Response({'error': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)

    permission = HasFirmAccess()
    if not permission.has_object_permission(request, None, bill.firm):
        return Response({'error': 'You do not have access to this invoice workspace.'}, status=status.HTTP_403_FORBIDDEN)

    bill.status = 'processing'
    bill.extraction_failed = False
    bill.validation_warnings = []
    bill.save()

    extract_invoice_data.delay(bill.id)

    log_audit(
        user=request.user,
        firm=bill.firm,
        resource_type=AuditLog.RESOURCE_BILL,
        resource_id=bill.id,
        action=AuditLog.ACTION_RETRY,
        request=request,
    )

    return Response({
        'id': bill.id,
        'status': bill.status,
        'message': 'AI extraction enqueued for retry.'
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_excel(request, firm_id):
    """
    POST: Export filtered/selected bills for a firm into a styled Excel workbook.
    Accepts bill_ids (list), or query filters start_date, end_date, status.
    """
    firm = get_firm_or_403(request, firm_id)
    if isinstance(firm, Response):
        return firm

    bills = Bill.objects.filter(firm=firm, is_deleted=False)

    # 1. Handle selection vs query-filters
    bill_ids = request.data.get('bill_ids')
    if isinstance(bill_ids, list):
        bills = bills.filter(id__in=bill_ids)
    else:
        status_filter = request.data.get('status')
        if status_filter:
            bills = bills.filter(status=status_filter)

        start_date = request.data.get('start_date')
        if start_date:
            bills = bills.filter(uploaded_at__date__gte=start_date)

        end_date = request.data.get('end_date')
        if end_date:
            bills = bills.filter(uploaded_at__date__lte=end_date)

        # Enforce verified only by default
        include_unverified = request.data.get('include_unverified', False)
        if not include_unverified:
            bills = bills.filter(status='verified')

    # Convert to list to track exported IDs
    bills_list = list(bills.order_by('-uploaded_at'))
    exported_bill_ids = [b.id for b in bills_list]

    if not exported_bill_ids:
        return Response({'error': 'No matching invoice records found for export.'}, status=status.HTTP_400_BAD_REQUEST)

    # 2. Group bills by upload date
    from collections import defaultdict
    bills_by_date = defaultdict(list)
    for b in bills_list:
        bills_by_date[b.uploaded_at.date()].append(b)

    sorted_dates = sorted(bills_by_date.keys())

    # 3. Create Spreadsheet Data
    import io

    import pandas as pd
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    def make_sheet_data(bills_sub):
        rows = []
        for b in bills_sub:
            r = b.raw_data or {}
            rows.append({
                'Date of Bill': r.get('invoice_date', ''),
                'Invoice Number': r.get('invoice_number', ''),
                'Seller (From)': r.get('party_name_from', ''),
                'Seller GSTIN': r.get('gstin_from', ''),
                'Buyer (To)': r.get('party_name_to', ''),
                'Buyer GSTIN': r.get('gstin_to', ''),
                'Place of Supply': r.get('place_of_supply', ''),
                'Taxable Amount': float(r.get('taxable_amount', 0.0) or 0.0),
                'CGST': float(r.get('cgst', 0.0) or 0.0),
                'SGST': float(r.get('sgst', 0.0) or 0.0),
                'IGST': float(r.get('igst', 0.0) or 0.0),
                'Total Amount': float(r.get('total_amount', 0.0) or 0.0),
            })
        return rows

    columns = [
        'Date of Bill', 'Invoice Number', 'Seller (From)', 'Seller GSTIN',
        'Buyer (To)', 'Buyer GSTIN', 'Place of Supply', 'Taxable Amount',
        'CGST', 'SGST', 'IGST', 'Total Amount'
    ]

    # 4. Generate openpyxl workbook
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for d in sorted_dates:
            sheet_name = d.strftime('%d-%m-%Y')
            sub_bills = bills_by_date[d]
            df = pd.DataFrame(make_sheet_data(sub_bills), columns=columns)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for d in sorted_dates:
            sheet_name = d.strftime('%d-%m-%Y')
            worksheet = workbook[sheet_name]
            worksheet.freeze_panes = 'A2'

            # Bold Header Fill & Font
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_idx in range(1, 13):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Auto-width calculation
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 13)

            # Currency styling
            currency_format = '"₹"#,##0.00'
            for row in range(2, worksheet.max_row + 1):
                for col_idx in [8, 9, 10, 11, 12]:
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = currency_format
                    cell.alignment = Alignment(horizontal='right')

            # Summary Totals Row
            total_row_idx = worksheet.max_row + 1
            total_font = Font(name='Calibri', size=11, bold=True)
            worksheet.cell(row=total_row_idx, column=1, value='Total').font = total_font

            border_style = Border(
                top=Side(style='thin', color='A0A0A0'),
                bottom=Side(style='double', color='1F4E79')
            )

            for col_idx in [8, 9, 10, 11, 12]:
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
                cell = worksheet.cell(row=total_row_idx, column=col_idx, value=formula)
                cell.font = total_font
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right')
                cell.border = border_style

    excel_data = output.getvalue()

    # 5. Upload file using helper service
    from datetime import datetime
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"invoices_{date_str}.xlsx"

    file_url = InvoiceStorageService.upload_export(excel_data, filename, firm.id)

    # 6. Save ExcelExportBatch record
    batch = ExcelExportBatch.objects.create(
        firm=firm,
        file_name=filename,
        file_url=file_url,
        exported_by=request.user,
        bill_ids=exported_bill_ids
    )

    # Create CloudVaultEntry
    from vault.models import CloudVaultEntry
    CloudVaultEntry.objects.create(
        firm=firm,
        excel_export=batch,
        file_name=filename,
        file_url=file_url,
        module='exports',
        is_finalized=True
    )

    log_audit(
        user=request.user,
        firm=firm,
        resource_type=AuditLog.RESOURCE_BILL,
        resource_id=batch.id,
        action=AuditLog.ACTION_EXPORT,
        details={'file_name': filename, 'exported_count': len(exported_bill_ids)},
        request=request,
    )

    return Response({
        'batch_id': batch.id,
        'file_name': batch.file_name,
        'file_url': batch.file_url,
        'exported_count': len(exported_bill_ids),
        'bill_ids': exported_bill_ids,
        'message': 'Excel export generated successfully.'
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated, HasFirmAccess])
def import_existing_excel(request, firm_id):
    """
    POST: Upload an existing LedgerPro Excel workbook.
    Reads each sheet named 'DD-MM-YYYY', parses rows, and creates verified Bill records
    with uploaded_at matching that sheet date.
    """
    firm = get_firm_or_403(request, firm_id)
    if isinstance(firm, Response):
        return firm

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return Response({'error': 'No excel file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    import re
    from datetime import datetime

    import pandas as pd
    from django.utils import timezone as django_timezone
    from django.utils.timezone import make_aware

    try:
        # Load the workbook
        xls = pd.ExcelFile(excel_file)
        imported_count = 0
        date_pattern = re.compile(r'^\d{2}-\d{2}-\d{4}$')

        # Column mapping helper
        def get_val(row_data, keys_list, default=''):
            for k in keys_list:
                if k in row_data and not pd.isna(row_data[k]):
                    return row_data[k]
            return default

        for sheet_name in xls.sheet_names:
            # Determine base sheet datetime
            aware_datetime = None
            if date_pattern.match(sheet_name):
                try:
                    sheet_date = datetime.strptime(sheet_name, '%d-%m-%Y')
                    aware_datetime = make_aware(sheet_date)
                except Exception:
                    pass

            if not aware_datetime:
                aware_datetime = django_timezone.now()

            df = xls.parse(sheet_name)
            # Normalize headers
            df.columns = [str(col).strip() for col in df.columns]

            for _, row in df.iterrows():
                # Skip summary total rows
                date_of_bill_val = get_val(row, ['Date of Bill', 'Date', 'Date Of Bill'])
                invoice_no_val = get_val(row, ['Invoice Number', 'Invoice No', 'Invoice number', 'Invoice no', 'Invoice No.'])

                if str(date_of_bill_val).strip().lower() == 'total' or pd.isna(invoice_no_val):
                    continue

                # Map columns
                invoice_number = str(invoice_no_val).split('.')[0].strip() # remove any float parsing decimals
                if not invoice_number or invoice_number == 'nan':
                    continue

                party_name_from = str(get_val(row, ['Seller (From)', 'Seller', 'From', 'Seller Name', 'Seller name'])).strip()
                party_name_to = str(get_val(row, ['Buyer (To)', 'Buyer', 'To', 'Buyer Name', 'Buyer name'])).strip()

                if party_name_from == 'nan' or party_name_from == '':
                    party_name_from = 'Unknown Seller'
                if party_name_to == 'nan' or party_name_to == '':
                    party_name_to = 'Unknown Buyer'

                bill_type = 'sale' if firm.name.lower() in party_name_from.lower() else 'purchase'

                # Try to parse 'Date of Bill' cell to use as uploaded_at
                date_of_bill_str = str(date_of_bill_val).strip()
                row_datetime = aware_datetime
                if date_of_bill_str and date_of_bill_str.lower() != 'nan':
                    # Try common date formats
                    for fmt in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            parsed_date = datetime.strptime(date_of_bill_str.split(' ')[0], fmt)
                            row_datetime = make_aware(parsed_date)
                            break
                        except Exception:
                            continue

                try:
                    taxable_amount = float(get_val(row, ['Taxable Amount', 'Assessable Amount', 'Taxable Value', 'Amount'], 0.0))
                except Exception:
                    taxable_amount = 0.0

                try:
                    cgst = float(get_val(row, ['CGST', 'Cgst', 'Cgst Amount', 'CGST Amount'], 0.0))
                except Exception:
                    cgst = 0.0

                try:
                    sgst = float(get_val(row, ['SGST', 'Sgst', 'Sgst Amount', 'SGST Amount'], 0.0))
                except Exception:
                    sgst = 0.0

                try:
                    igst = float(get_val(row, ['IGST', 'Igst', 'Igst Amount', 'IGST Amount'], 0.0))
                except Exception:
                    igst = 0.0

                try:
                    total_amount = float(get_val(row, ['Total Amount', 'Total Amt', 'Total', 'Total Value'], 0.0))
                except Exception:
                    total_amount = 0.0

                raw_data = {
                    'bill_type': bill_type,
                    'invoice_date': date_of_bill_str if date_of_bill_str != 'nan' else row_datetime.strftime('%Y-%m-%d'),
                    'invoice_number': invoice_number,
                    'party_name_from': party_name_from,
                    'gstin_from': str(get_val(row, ['Seller GSTIN', 'Seller GST', 'Seller Gstin', 'Seller Gst'])).strip() if not pd.isna(get_val(row, ['Seller GSTIN', 'Seller GST', 'Seller Gstin', 'Seller Gst'])) else '',
                    'party_name_to': party_name_to,
                    'gstin_to': str(get_val(row, ['Buyer GSTIN', 'Buyer GST', 'Buyer Gstin', 'Buyer Gst'])).strip() if not pd.isna(get_val(row, ['Buyer GSTIN', 'Buyer GST', 'Buyer Gstin', 'Buyer Gst'])) else '',
                    'place_of_supply': str(get_val(row, ['Place of Supply', 'Place Of Supply', 'Pos', 'POS'])).strip() if not pd.isna(get_val(row, ['Place of Supply', 'Place Of Supply', 'Pos', 'POS'])) else '',
                    'taxable_amount': taxable_amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst,
                    'total_amount': total_amount,
                }

                # Create Bill record
                Bill.objects.create(
                    firm=firm,
                    status='verified',
                    file_name=f"Imported_{sheet_name}_{invoice_number}.pdf",
                    file_url='',
                    file_size=0,
                    uploaded_by=request.user,
                    uploaded_at=row_datetime,
                    raw_data=raw_data
                )
                imported_count += 1

        return Response({
            'message': f'Excel file imported successfully. Parsed {imported_count} records.',
            'imported_count': imported_count
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': f'Failed to parse excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)



